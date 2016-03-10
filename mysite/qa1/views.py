from django.http import HttpResponse
from django.template import loader
from django.shortcuts import render
from django.http import Http404
from django.shortcuts import get_object_or_404
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
from django.shortcuts import redirect
from django.contrib.auth import logout as django_logout
from django.contrib.auth.decorators import login_required
# from .forms import Mcq_Question_Form
# from .forms import ExcelForm
from .models import Mcq_Question
# from .forms import *
from django.contrib.auth.models import User
import openpyxl
from django.contrib.admin.views.decorators import staff_member_required
from readingmaterial.models import *

from question.forms import *

from question.models import *
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.models import Permission
from django.contrib.auth.decorators import permission_required

import time

from django.contrib.auth import authenticate, login

def index(request):
    return HttpResponse("Hello, world. You're at the ....qa...... index.")




def detail(request):
    return HttpResponse("Hello, world. You're at the qa..........details.... index.")


def mylogin(request):
    if request.method == "POST":
        uname = request.POST.get('username', "")
        pword = request.POST.get('password',"")

        user = authenticate(username= uname, password= pword)

        print ("******** trying to authenticate user with ")
        print (uname)
        print (pword)
        print (user)
        if user is not None:
            login(request, user)
            print ("redircting to dashboard")
            return redirect('/dashboard/')
        else:
            return redirect('/login/')


    return render(request, 'readingmaterial/registration/login_new.html')









def mylogout(request):
    print ("\n\ngoint to logout the user")
    django_logout(request)
    return redirect('/login/')




def register(request):

    print ("**** in register method\n\n\n")

    if request.method == "POST":
        username = request.POST.get('username', "")
        fname = request.POST.get('first_name',"")
        lname = request.POST.get('last_name', "")
        email = request.POST.get('email',"")
        password = request.POST.get('password',"")

        if (username == 'User Name' or password=="Password"):
            return HttpResponse("Pls Fill Up Another User Name and Password")
        

        # u = str(username)
        # print ("\n******  user name is: " + u)
        # print (u)
        # print (type(u))


        # user = User.objects.get(username = u)

        # if (user is  None):
        #     return HttpResponse("Pls Try Another User Name ")




        if (username and fname and lname and email and password):

            u = User.objects.create_user(username, email, password, first_name=fname, last_name=lname)
            u.save()
            # return HttpResponse("Registration complete! ")
            return redirect("/login/") 
        else:
            return HttpResponse("Pls Fill Up All the Fields perfectly Or try a new username")


    return render(request, 'readingmaterial/registration/register.html')

















# @user_passes_test(can_upload_mcq_from_excel)

# @user_passes_test(can_upload_mcq_from_excel, login_url='/login/')
# @permission_required('question.can_upload_mcq_from_excel')
@staff_member_required
@permission_required('question.excel_upload', login_url='/admin/login/')
def excelparsing(request):

    timestamp = int(time.time())



    if request.method == 'POST':
        form = ExcelForm2(request.POST, request.FILES)
        print ("\n this is a post method \n")
        tag = request.POST.get('tag', '')
        if (not tag):
            tag = str(timestamp)
        if form.is_valid():
            print ("the form is valid\n")
            wb = openpyxl.load_workbook(request.FILES['excel_file'])
            sheet = wb.get_sheet_by_name('Sheet1')
            print ("\n\n type of sheet is: " + str(type(sheet)))
            for row in range(1, 4):
                if ((sheet.cell(row=row, column=1).value) is None):
                    print ("in if part: " + str(sheet.cell(row=row, column=1).value))
                    break

                q = Mcq_Question(question_text=str(sheet.cell(row=row, column=1).value))
                #q = Mcq_Question(question_text="alamin is hungry now")
                q.save()
                c1 = str(sheet.cell(row=row, column=2).value)
                c2 = str(sheet.cell(row=row, column=3).value)
                c3 = str(sheet.cell(row=row, column=4).value)
                c4 = str(sheet.cell(row=row, column=5).value)

                a = str(sheet.cell(row=row, column=6).value)

                t1 = str(sheet.cell(row=row, column=7).value)
                t2 = str(sheet.cell(row=row, column=8).value)

                q.choice_a = c1
                q.choice_b = c2
                q.choice_c = c3
                q.choice_d = c4

                q.mcq_answer = a
                q.tag1 = t1
                q.tag2 = t2
                #q.tag3 = t3
                q.tag5 = tag
                q.update_date()

                q.save()

                t3 = str(sheet.cell(row=row, column=9).value)


                






            	# for col in range(1, 10):
            	# 	print (row, col, str(sheet.cell(row=row, column=col).value))
            		
            	# 	if ()

           
           
            return HttpResponse('excel has been uploaded successfully')

        else:
        	return HttpResponse("......................failed to upload the ExcelForm")
    else:
        print ('\nthe form is not valid so returning\n')
        form = ExcelForm2()
        tag = "#" + str(timestamp) + "#"
        # form.save()
    
    return render(request, 'question/excelparsing.html', {'form': form, 'tag': tag})










def fileupload(request):
    HttpResponse("in fileupload view. this view needs to be edited")
    # if request.method == 'POST':
    #     form = ImageForm(request.POST, request.FILES)
    #     print ("\n this is a post method \n")
    #     if form.is_valid():
    #         print ("the form is valid\n")
    #         form.save()
           
    #         return HttpResponse('Image has been uploaded successfully')

    #     else:
    #       return HttpResponse("......................failed to upload the image")
    # else:
    #     print ('\nthe form is not valid so returning\n')
    #     form = ImageForm()
    
    # return render(request, 'exam/fileupload.html', {'form': form})

