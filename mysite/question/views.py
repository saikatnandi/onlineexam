from django.http import HttpResponse
from django.template import loader
from django.shortcuts import render
from django.http import Http404
from django.shortcuts import get_object_or_404
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
from django.shortcuts import redirect
from django.contrib.auth import logout as django_logout
from django.contrib.auth.decorators import login_required

from .forms import *
from django.contrib.auth.models import User
import openpyxl
from  qa1.models import *
from readingmaterial.models import *
from django.shortcuts import render_to_response
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponseRedirect
from  django.template.context_processors import csrf

from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.models import Permission
from django.contrib.auth.decorators import permission_required

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from readingmaterial.models import *
from subscription.models import *

import time
import random
from django.db.models import Q

from .views_1 import *









@staff_member_required
@permission_required('question.excel_upload', login_url='/admin/login/')
def create_question_topic_wise(request):
    form = create_question_topic_wise_form()

    # print ("\n***** create_question_topic_wise")
    if request.method == 'POST':
    	# print ("\n***** form is submitted by post\n")
        form = create_question_topic_wise_form(request.POST)
        # print ("\n this is a post method \n")

        if form.is_valid():
            question_topic = form.cleaned_data.get('question_topic', "")

            # print ("****** type of question_topic: ")
            # print (type(question_topic))

            # if (question_topic):
            #     question_topic = Question_Topic.objects.filter(estion_topic_text = question_topic)
            #     question_topic = question_topic[0]




            question_title = form.cleaned_data.get('question_title')
            total_question = form.cleaned_data.get('total_question')
            marks = form.cleaned_data.get('marks')
            negative_marks = form.cleaned_data.get('negative_marks')
            topic_list = form.cleaned_data.get('topic_list')

            all_mcq = []
            if (topic_list):
                # tl = topic_list[1]
                # print ("****type of t1 is:\n")
                # print (type(tl))
                all_mcq = Mcq_Question.objects.filter(readingcontent__reading_topic__in = topic_list)
            # print (tl)
            # mcq = []
            mcq_set = set()
            if (len(all_mcq) <= total_question):

                mcq_set = set(all_mcq)
            else:
                all_mcq_len = len(all_mcq)
                while (len(mcq_set) < total_question):
                    print ("... in while loop adding mcq ...")
                    ran = random.randint(0, all_mcq_len-1)
                    mcq_set.add(all_mcq[ran])




            print ("******** the selected mcqs are  ***\n\n")
            print (mcq_set)




            # if (question_topic):
            #     question_topic = Question_Topic.objects.filter(estion_topic_text = question_topic)
            #     question_topic = question_topic[0]

            question_set = Question_Set(question_set_text = question_title)
            question_set.uploader = request.user
            question_set.question_topic = question_topic


            question_set.individual_mcq_marks = marks
            question_set.negative_marking_percentage = negative_marks
            question_set.save()

            for mcq in mcq_set:
                question_set.mcq_question.add(mcq)

            question_set.update_date()
            question_set.save()






            # print ("*********** title is: " + question_title)

            # print ("\n\n****** form question_topic " + str(question_topic))
            # print (topic_list)
            # for field in form:
            #     print ("*******form.field..")
            #     print (field)

            return HttpResponse('form submission is  successful')

        # else:
        # 	return HttpResponse("......................form submission is failed ")





    c = {}
    c.update(csrf(request))
    c.update({'form':form})
    # return render(request, 'question/create_question_topic_wise.html', {'form': form})
    return render_to_response ('question/create_question_topic_wise.html', c)



















@staff_member_required
# @permission_required('question.excel_upload', login_url='/admin/login/')
def create_question_sub_topic_wise(request):
    form = create_question_sub_topic_wise_form()
    # print ("\n***** create_question_sub_topic_wise")
    if request.method == 'POST':
        print ("\n***** create_question****sub_topic_wise")
        form = create_question_sub_topic_wise_form(request.POST)

        if form.is_valid():  
            for field in form:
                print ("*******form.field..")
                print (field)

            return HttpResponse('form submission is  successful')



    c = {}
    c.update(csrf(request))
    c.update({'form':form})

    return render(request, 'question/create_question_sub_topic_wise.html', {'form': form})







def create_mcq_from_excel(request, sheet, row, tag):

    if ((sheet.cell(row=row, column=1).value) is None):
        # print ("in if part: " + str(sheet.cell(row=row, column=1).value))
        return None
    # print ('***********.....mcq title is about to be been created \n')
    title = (sheet.cell(row=row, column=1).value)
    if (title):

        try:
            title = str(title)
        except Exception:
            title = title.encode('utf8')

        # title = title.encode('utf8')

        # print ("title is: ")
        # print (title)
        q = Mcq_Question(question_text=title)
    # print ('***********.....mcq title has been created \n')
    #q = Mcq_Question(question_text="alamin is hungry now")
    q.save()



    c1 = (sheet.cell(row=row, column=2).value)
    if (c1):
        # print (type(c1))
        # c1 =str(c1)
        try:
            c1 = str(c1)
        except Exception:
            c1 = c1.encode('utf8')


    c2 = (sheet.cell(row=row, column=3).value)
    if(c2):

        try:
            c2 = str(c2)
        except Exception:
            c2 = c2.encode('utf8')

        # c2 = c2.encode('utf8')

    c3 = (sheet.cell(row=row, column=4).value)
    if (c3):

        try:
            c3 = str(c3)
        except Exception:
            c3 = c3.encode('utf8')


        # c3 = c3.encode('utf8')

    c4 = (sheet.cell(row=row, column=5).value)
    if(c4):

        try:
            c4 = str(c4)
        except Exception:
            c4 = c4.encode('utf8')

        # c4 = c4.encode('utf8')



    c5 = (sheet.cell(row=row, column=8).value)
    if (c5):

        try:
            c5 = str(c5)
        except Exception:
            c5 = c5.encode('utf8')

        # c5 = c5.encode('utf8')

    c6 = (sheet.cell(row=row, column=9).value)
    if (c6):

        try:
            c6 = str(c6)
        except Exception:
            c6 = c6.encode('utf8')

        # c6 = c6.encode('utf8')



    # c4 = (sheet.cell(row=row, column=5).value).encode('utf8')

    # c5 = (sheet.cell(row=row, column=8).value).encode('utf8')
    # c6 = (sheet.cell(row=row, column=9).value).encode('utf8')



    a = (sheet.cell(row=row, column=6).value)

    if (a):

        try:
            a = str(a)
        except Exception:
            a = a.encode('utf8')

        # a = a.encode('utf8')

    explanation = (sheet.cell(row=row, column=7).value)
    if (explanation):

        try:
            explanation = str(explanation)
        except Exception:
            explanation = explanation.encode('utf8')

        # explanation = explanation.encode('utf8')


    t1 = (sheet.cell(row=row, column=10).value)
    if (t1):
        try:
            t1 = str(t1)
        except Exception:
            t1 = t1.encode('utf8')

        # t1 = t1.encode('utf8')
    t2 = (sheet.cell(row=row, column=11).value)
    if (t2):
        try:
            t2 = str(t2)
        except Exception:
            t2 = t2.encode('utf8')

        # t2 = t2.encode('utf8')
    t3 = (sheet.cell(row=row, column=12).value)
    if (t3):
        try:
            t3 = str(t3)
        except Exception:
            t3 = t3.encode('utf8')

        # t3= t3.encode('utf8')
    t4 = (sheet.cell(row=row, column=13).value)
    if (t4):
        try:
            t4 = str(t4)
        except Exception:
            t4 = t4.encode('utf8')

        # t4= t4.encode('utf8')


    q.choice_a = c1
    q.choice_b = c2
    q.choice_c = c3
    q.choice_d = c4

    # q.choice_e = c5
    # q.choice_f = c6




    q.mcq_answer = a
    q.explanation_text = explanation

    q.tag1 = t1
    q.tag2 = t2
    q.tag3 = t3
    q.tag5 = tag
    q.update_date()
    if (not q.uploader):
        q.uploader = request.user

    q.save()

    print ('***********.....mcq has been uploaded and saved returning\n')
    # print (q)
    print ()

    return q





@staff_member_required
@permission_required('question.excel_upload', login_url='/admin/login/')
def excelparsing(request):

    timestamp = int(time.time())
    content = ""
    sub_topic = ""
    topic = ""

    if request.method == 'POST':
        form = ExcelForm2(request.POST, request.FILES)
        print ("\n this is a post method \n")
        tag = request.POST.get('tag', '')
        if (not tag):
            tag = str(timestamp)
        if form.is_valid():
            content = form.cleaned_data.get('content', "")
            if (content):
                content = ReadingContent.objects.get(content_title = str(content))

            if (content):
                sub_topic = SubTopic1.objects.get(readingcontent = content)
                if (sub_topic):
                    topic = ReadingTopic.objects.get(subtopic1 = sub_topic)



            print ("the form is valid\n")
            wb = openpyxl.load_workbook(request.FILES['excel_file'])
            sheet = wb.get_sheet_by_name('Sheet1')
            print ("\n\n type of sheet is: " + str(type(sheet)))
            for row in range(10, 10000):

                if ((sheet.cell(row=row, column=1).value) is None):
                    print ("in if part: " + str(sheet.cell(row=row, column=1).value))
                    break

                q = create_mcq_from_excel(request, sheet, row, tag)
                if (q is None):
                    break





                if (content):
                    q.tag_content = str(content)
                    content.mcq_question.add(q)

                    if (sub_topic):
                        q.tag_sub_topic = str(sub_topic)
                        # content.subtopic1 = sub_topic

                        if (topic):
                            q.tag_topic = str(topic)
                            # content.topic = topic



                    
                    

                q.save()


                if (content):
                    content.save()


           
           
            return HttpResponse('excel has been uploaded successfully')

        else:
        	return HttpResponse("......................failed to upload the ExcelForm")
    else:
        print ('\nthe form is not valid so returning\n')
        form = ExcelForm2()
        tag = "#" + str(timestamp) + "#"
        # form.save()
    
    return render(request, 'question/excelparsing.html', {'form': form, 'tag': tag})







def create_quick_question_from_excel(request, sheet, row,reading_content, tag):

    if ((sheet.cell(row=row, column=1).value) is None):
        # print ("in if part: " + str(sheet.cell(row=row, column=1).value))
        return None
    # print ('***********.....mcq title is about to be been created \n')
    question_text = (sheet.cell(row=row, column=1).value)
    if (question):
        question_text = getString(question_text)

        q = Quick_Question(quick_question_text=question_text)
    # print ('***********.....mcq title has been created \n')
    #q = Mcq_Question(question_text="alamin is hungry now")
    q.content = reading_content
    # q.save()
    # q.content = reading_content
  



    answer = (sheet.cell(row=row, column=3).value)
    if (answer):
        answer = getString(answer)
        q.quick_question_answer = answer





    # q.update_date()
    # if (not q.uploader):
    #     q.uploader = request.user

    # q.save()



    return q
















@staff_member_required
@permission_required('question.excel_upload', login_url='/admin/login/')
def excel_quickquestion(request):
    form =  Excel_quickquestion_form()
    # timestamp = int(time.time())
    content = ""
    # sub_topic = ""
    # topic = ""

    quick_question_list = []

    if request.method == 'POST':
        # print ("\n***** form is submitted by post\n")
        form =  Excel_quickquestion_form (request.POST, request.FILES)
        # print ("\n this is a post method \n")

        print ("\n\n************going to check if form is valid")

        if form.is_valid():
            print ("\n******** valid\n\n")
            reading_content = form.cleaned_data.get('reading_content', "")
            print (type(reading_content))
            # reading_content = ReadingContent.objects.get(content_title=reading_content)
            # print (type(reading_content))
            print ("********* reading content value\n\n")
            print(reading_content)
            tag = form.cleaned_data.get('tag',"")


            wb = openpyxl.load_workbook(request.FILES['excel_file'])
            sheet = wb.get_sheet_by_name('Sheet1')
            # print ("\n\n type of sheet is: " + str(type(sheet)))
            for row in range(10, 10000):
                if ((sheet.cell(row=row, column=1).value) is None):
                    # print ("in if part: " + str(sheet.cell(row=row, column=1).value))
                    break

                q = create_quick_question_from_excel(request, sheet, row, reading_content, tag)
                if (q is None):
                    break     

                quick_question_list.append(q)   


            for q in quick_question_list:
                q.save()    
                q.update_date()
                if (not q.uploader):
                    q.uploader = request.user

                q.save()
                # q.save()

            return HttpResponse('form submission is  successful')

        else:
            return HttpResponse('f********* NOT VALID ******')

        # else:
        #   return HttpResponse("......................form submission is failed ")


    


    c = {}
    c.update(csrf(request))
    # form = Excel_quickquestion_form()
    c.update({'form':form})
    # return render(request, 'question/create_question_topic_wise.html', {'form': form})
    # form =  Excel_quickquestion_form()
    return render_to_response ('question/excel_quickquestion.html', c)



















@staff_member_required
@permission_required('question.excel_upload', login_url='/admin/login/')
def excelparsing_question_set(request):

    timestamp = int(time.time())


    if request.method == 'POST':
        form = Upload_Question_Set_From_Excel_Form(request.POST, request.FILES)
        print ("\n this is a post method \n")
        tag = request.POST.get('tag', '')
        if (not tag):
            tag = str(timestamp)
        if form.is_valid():
            question_set = form.cleaned_data.get('question_set', "")
            # question_set = Question_Set.objects.filter(question_set_text = str(question_set))
            
            if (question_set):
                pass
                # question_set = question_set[0]
                # marks = form.cleaned_data.get('marks')
                # negative_marks = form.cleaned_data.get('negative_marks')











            else:
                print ("******************* No question Set found\n")
                return
            


            print ("the form is valid\n")
            wb = openpyxl.load_workbook(request.FILES['excel_file'])
            sheet = wb.get_sheet_by_name('Sheet1')
            print ("\n\n type of sheet is: " + str(type(sheet)))
            for row in range(10, 10000):
                if ((sheet.cell(row=row, column=1).value) is None):
                    print ("in if part: " + str(sheet.cell(row=row, column=1).value))
                    break

                q = create_mcq_from_excel(request, sheet, row, tag)





                if (question_set):
                    # q.tag_content = str(content)
                    question_set.mcq_question.add(q)

                


                q.save()

                # question_set.individual_mcq_marks = marks
                # question_set.negative_marking_percentage = negative_marks  
                question_set.save()

                t3 = str(sheet.cell(row=row, column=9).value)

           
           
            return HttpResponse('**********Question_Set ******* has been uploaded successfully')

        else:
            return HttpResponse("......................failed to upload the ExcelForm")
    else:
        print ('\nthe form is not valid so returning\n')
        form = Upload_Question_Set_From_Excel_Form()
        tag = "#" + str(timestamp) + "#"
        # form.save()
    
    return render(request, 'question/excelparsing_question_set.html', {'form': form, 'tag': tag})













